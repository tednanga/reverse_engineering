import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
import string
import itertools
import re
import time

########################################################################
################ creation data for extraction Diagnostic ###############
########################################################################
columns = list(string.ascii_uppercase)

# covers mortalité < P5
#----------------------------
ranges = [
    ('Diagnostic mortalité', [
        ('A', 'J'),
        ('Q', 'Z'),
        ('AG', 'AP'), 
        ('AW', 'BF') # P4
        ]),
    ('Diagnostic mortalité après pesée', [
        ('K', 'P'), 
        ('AA', 'AF'), 
        ('AQ', 'AV'), 
        ('BG', 'BL') # P4
        ])
    ]

list_cols_A = columns + ['AA'] + [
    ''.join((item[0], item[1])) 
    for item in itertools.combinations(columns, 2) 
    if item[0].startswith('A')
]

list_cols_B = ['BA', 'BB'] + [
    ''.join((item[0], item[1])) 
    for item in itertools.combinations(columns, 2) 
    if item[0].startswith('B')
    and len(item) > 1
]

list_cols = list_cols_A + list_cols_B

dict_colPeriod = {}

for idx in range(len(ranges)):
    typ = ranges[idx]
    for i in range(len(typ[1])):
        idx_l, idx_r = list_cols.index(typ[1][i][0]), list_cols.index(typ[1][i][1])
        dict_colPeriod.update(
            {
                k: typ[0] + f' P_{i + 1}' 
                for k in list_cols[idx_l:idx_r + 1]
            }
        )
        

###################################################################
################ function for extraction Diagnostic ###############
###################################################################
def extract_diagnosis(input_source, comments=False):
    import os
    
    if isinstance(input_source, pd.ExcelFile):
        path_file = input_source._io  # Recover file path from ExcelFile
    else:
        path_file = input_source
    
    def time_info(start):
        temp = time.time() - start
        hours = temp // 3600
        temp -= 3600 * hours
        minutes = temp // 60
        seconds = temp - 60 * minutes
        print('extract Diagnostic in: %d:%d:%d' % (hours, minutes, seconds))
    
    if comments:
        print('load cell comments file {}'.format(os.path.basename(path_file)))
        start = time.time()
        wb = load_workbook(path_file)
        time_info(start)
    else:
        wb = load_workbook(path_file)
    
    mort = wb['mortalité']
    
    comments = {
        str(cell).split('.')[1][:-1]: cell.comment.text
        for row in mort.iter_rows()
        for cell in row
        if cell.comment
    }

    cols_unique = np.unique([re.sub('[0-9]', '', k) for k in comments.keys()]).tolist()
    
    frames = {}
    if len(cols_unique) > 0:
        for col in cols_unique:
            Ser = pd.Series({
                int(re.findall(r'\d+', k)[0]) - 3: v.split(':\n')[-1].strip()
                for k, v in comments.items()
                if col == re.sub('[0-9]', '', k)
            })
            Ser.name = 'Diagnostic'
            key = '_'.join(dict_colPeriod[col].split(' ')[1:-1])
            D_frame = Ser.to_frame(Ser.name).reset_index()
            D_frame['period'] = dict_colPeriod[col].split(' ')[-1]
            if key not in frames:
                frames[key] = [D_frame]
            else:
                frames[key].append(D_frame)

    return {k: pd.concat(v, axis=0, ignore_index=False) for k, v in frames.items()}
